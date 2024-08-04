import openpyxl as op
from datetime import datetime
import json


class JsonWorker:
    def __init__(self):
        self.file_name = 'fuel_data'

    def convert_xlsx_to_json(self: str):
        wb = op.load_workbook(self)
        sheet = wb.active
        rows_cnt = sheet.max_row
        record_dict = {}
        for i in range(6, rows_cnt + 1):
            if sheet.cell(i, 1).value:
                date_str = str(sheet.cell(i, 1).value)
                date_obj = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
                date_to_append = str(date_obj.date())

                dict_data = {
                    'date': date_to_append,
                    'sum': sheet.cell(i, 2).value,
                    'liter': sheet.cell(i, 3).value,
                    'type': 'LPG' if sheet.cell(i, 5).value == 'Г' else 'Gasoline',
                    'km': sheet.cell(i, 6).value, }
                dict_to_append = {i: dict_data}
                record_dict.update(dict_to_append)
        return record_dict

    def add_item_to_json(self, json_data, summ: str, volume: str, fuel_type: str, milage: str, date_add):
        dict_data = {
            'sum': summ,
            'liter': volume,
            'type': fuel_type,
            'km': milage, }

        date_str = str(date_add)
        print(date_str)
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        date_to_append = str(date_obj.date())

        dict_to_append = {date_to_append: dict_data}
        json_data.update(dict_to_append)
        return json_data

    def save_json_to_file(self, data):
        with open(self.file_name, 'w') as f:
            json.dump(data, f)
