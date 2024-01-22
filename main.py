import openpyxl as op
from datetime import datetime
import json

filename = 'fuel_data.xlsx'
wb = op.load_workbook(filename)
sheet = wb.active
rows_cnt = sheet.max_row
record_dict = {}

for i in range(6, rows_cnt + 1):
    if sheet.cell(i, 1).value:
        date_str = str(sheet.cell(i, 1).value)
        date_obj = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
        date_to_append = str(date_obj.date())

        dict_data = {
            'sum': sheet.cell(i, 2).value,
            'liter': sheet.cell(i, 3).value,
            'type': 'LPG' if sheet.cell(i, 5).value == 'Г' else 'Gasoline',
            'km': sheet.cell(i, 6).value, }
        dict_to_append = {date_to_append: dict_data}
        record_dict.update(dict_to_append)

with open('fuel_data.json', 'w') as f:
    json.dump(record_dict, f)
